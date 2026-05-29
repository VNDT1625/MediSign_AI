#!/usr/bin/env python3
"""Nạp ngược các chỉnh sửa của BÁC SĨ (file .xlsx) vào JSON gốc.

Đọc file Excel do `export_for_doctor_review.py` tạo ra (sau khi bác sĩ đã sửa),
ghép theo cột "ID (không sửa)" và cập nhật lại các file JSON trong knowledge_base.

An toàn:
  - Mặc định chạy ở chế độ DRY-RUN (chỉ in ra thay đổi, KHÔNG ghi file).
  - Thêm cờ --apply để thực sự ghi. Khi ghi, tự backup file gốc sang .bak.<timestamp>.
  - Chỉ cập nhật bản ghi có "Trạng thái duyệt" = 'Đã duyệt - đúng' / 'Đã sửa' / 'Cần xóa'.
  - Bản ghi 'Cần xóa' được đánh dấu needs_deletion=True (không xóa cứng).

Cách chạy:
    # Xem trước thay đổi (không ghi):
    python scripts/qa/import_doctor_review.py --in data/review/medisign_doctor_review.xlsx

    # Ghi thật (có backup tự động):
    python scripts/qa/import_doctor_review.py --in data/review/medisign_doctor_review.xlsx --apply
"""
from __future__ import annotations

import argparse
import json
import shutil
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
KB = ROOT / "data" / "knowledge_base"

DISEASE_FILES = [
    KB / "vietnam_common_diseases.json",
    KB / "vietnam_diseases_full.json",
]
INTERACTION_FILE = KB / "drug_interactions.json"

SEVERITY_FROM_VI = {
    "nặng (high)": "high",
    "trung bình (medium)": "medium",
    "nhẹ (low)": "low",
    "high": "high",
    "medium": "medium",
    "low": "low",
}
APPLY_STATUSES = {"đã duyệt - đúng", "đã sửa", "cần xóa", "cần xem lại"}


def _split_lines(text: Any) -> list[str]:
    if text is None:
        return []
    return [ln.strip() for ln in str(text).splitlines() if ln.strip()]


def _norm(value: Any) -> str:
    return str(value or "").strip().lower()


def _read_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {headers[i]: row[i] for i in range(len(headers))}
        if _norm(record.get("ID (không sửa)")):
            out.append(record)
    return out


def _load_array(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [x for x in data if isinstance(x, dict)]


def _apply_disease_edits(record: dict[str, Any], edit: dict[str, Any]) -> list[str]:
    """Trả về danh sách mô tả thay đổi."""
    changes: list[str] = []
    s = record.setdefault("structured", {})

    def set_list(field: str, new_text: Any, label: str) -> None:
        new_list = _split_lines(new_text)
        if new_list and new_list != (s.get(field) or []):
            changes.append(f"{label}: {len(s.get(field) or [])}→{len(new_list)} mục")
            s[field] = new_list

    sym_field = "common_symptoms" if "common_symptoms" in s or not s.get("symptoms") else "symptoms"
    set_list(sym_field, edit.get("Triệu chứng thường gặp (sửa được)"), "triệu chứng")
    set_list("red_flags", edit.get("Dấu hiệu cần khám gấp / red flags (sửa được)"), "red_flags")
    set_list("common_complications", edit.get("Biến chứng (sửa được)"), "biến chứng")

    new_sev = SEVERITY_FROM_VI.get(_norm(edit.get("Mức độ (sửa được)")))
    if new_sev and new_sev != _norm(s.get("severity")):
        changes.append(f"mức độ: {s.get('severity') or '∅'}→{new_sev}")
        s["severity"] = new_sev

    advice = edit.get("Lời khuyên / xử trí (sửa được)")
    if advice and str(advice).strip() and str(advice).strip() != str(s.get("advice") or "").strip():
        s["advice"] = str(advice).strip()
        changes.append("lời khuyên: cập nhật")

    content = edit.get("Mô tả đầy đủ (sửa được)")
    if content and str(content).strip() and str(content).strip() != str(record.get("content") or "").strip():
        record["content"] = str(content).strip()
        changes.append("mô tả: cập nhật")

    return changes


def _apply_interaction_edits(record: dict[str, Any], edit: dict[str, Any]) -> list[str]:
    changes: list[str] = []
    s = record.setdefault("structured", {})

    new_sev = SEVERITY_FROM_VI.get(_norm(edit.get("Mức độ (sửa được)")))
    if new_sev and new_sev != _norm(s.get("severity")):
        changes.append(f"mức độ: {s.get('severity') or '∅'}→{new_sev}")
        s["severity"] = new_sev

    for field, col, label in [
        ("mechanism", "Cơ chế tương tác (sửa được)", "cơ chế"),
        ("recommendation", "Khuyến nghị xử trí (sửa được)", "khuyến nghị"),
    ]:
        val = edit.get(col)
        if val and str(val).strip() and str(val).strip() != str(s.get(field) or "").strip():
            s[field] = str(val).strip()
            changes.append(f"{label}: cập nhật")

    content = edit.get("Mô tả đầy đủ (sửa được)")
    if content and str(content).strip() and str(content).strip() != str(record.get("content") or "").strip():
        record["content"] = str(content).strip()
        changes.append("mô tả: cập nhật")

    return changes


def _mark_reviewed(record: dict[str, Any], status: str, note: Any) -> None:
    record["needs_medical_review"] = status == "cần xem lại"
    record["doctor_reviewed"] = True
    record["doctor_review_status"] = status
    if status == "cần xóa":
        record["needs_deletion"] = True
    if note and str(note).strip():
        record["doctor_note"] = str(note).strip()
    if status in {"đã duyệt - đúng", "đã sửa"}:
        record["confidence"] = "high"


def process_file(path: Path, edits_by_id: dict[str, dict[str, Any]],
                 apply_fn, apply: bool) -> dict[str, int]:
    stats = {"matched": 0, "changed": 0, "marked": 0}
    records = _load_array(path)
    if not records:
        return stats

    dirty = False
    for rec in records:
        rid = _norm(rec.get("id"))
        edit = edits_by_id.get(rid)
        if not edit:
            continue
        status = _norm(edit.get("✅ Trạng thái duyệt"))
        if status not in APPLY_STATUSES:
            continue
        stats["matched"] += 1

        field_changes: list[str] = []
        if status in {"đã sửa", "đã duyệt - đúng", "cần xem lại"}:
            field_changes = apply_fn(rec, edit)
        _mark_reviewed(rec, status, edit.get("📝 Ghi chú của bác sĩ"))
        stats["marked"] += 1
        if field_changes:
            stats["changed"] += 1
            dirty = True
            print(f"  [{rec.get('id')}] {status} | " + "; ".join(field_changes))
        else:
            dirty = True  # mark fields changed too
            print(f"  [{rec.get('id')}] {status} (chỉ gắn nhãn duyệt)")

    if apply and dirty:
        backup = path.with_suffix(path.suffix + f".bak.{int(time.time())}")
        shutil.copy2(path, backup)
        path.write_text(
            json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"  💾 Đã ghi {path.name} (backup: {backup.name})")
    return stats


def main() -> None:
    ap = argparse.ArgumentParser(description="Nạp ngược chỉnh sửa của bác sĩ vào JSON")
    ap.add_argument("--in", dest="inp",
                    default=str(ROOT / "data" / "review" / "medisign_doctor_review.xlsx"))
    ap.add_argument("--apply", action="store_true",
                    help="Ghi thật vào file JSON (mặc định chỉ DRY-RUN).")
    args = ap.parse_args()

    xlsx = Path(args.inp)
    if not xlsx.exists():
        raise SystemExit(f"❌ Không tìm thấy file: {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    disease_edits: dict[str, dict[str, Any]] = {}
    interaction_edits: dict[str, dict[str, Any]] = {}
    if "Bệnh" in wb.sheetnames:
        for r in _read_sheet(wb["Bệnh"]):
            disease_edits[_norm(r.get("ID (không sửa)"))] = r
    if "Tương tác thuốc" in wb.sheetnames:
        for r in _read_sheet(wb["Tương tác thuốc"]):
            interaction_edits[_norm(r.get("ID (không sửa)"))] = r

    mode = "GHI THẬT (--apply)" if args.apply else "DRY-RUN (chỉ xem trước)"
    print(f"Chế độ: {mode}\n")

    print("── BỆNH ──")
    d_total = {"matched": 0, "changed": 0, "marked": 0}
    for path in DISEASE_FILES:
        st = process_file(path, disease_edits, _apply_disease_edits, args.apply)
        for k in d_total:
            d_total[k] += st[k]

    print("\n── TƯƠNG TÁC THUỐC ──")
    i_st = process_file(INTERACTION_FILE, interaction_edits, _apply_interaction_edits, args.apply)

    print("\n── TỔNG KẾT ──")
    print(f"  Bệnh        : khớp {d_total['matched']}, có sửa nội dung {d_total['changed']}, gắn nhãn {d_total['marked']}")
    print(f"  Tương tác   : khớp {i_st['matched']}, có sửa nội dung {i_st['changed']}, gắn nhãn {i_st['marked']}")
    if not args.apply:
        print("\n⚠️  Đây là DRY-RUN. Thêm --apply để ghi thật (sẽ tự backup).")


if __name__ == "__main__":
    main()
