#!/usr/bin/env python3
"""Xuất dữ liệu Bệnh + Tương tác thuốc ra file Excel cho BÁC SĨ rà soát.

Mục tiêu: tạo 1 file .xlsx trực quan để bác sĩ:
  - Đọc đầy đủ thông tin từng bệnh / từng tương tác thuốc.
  - Sửa trực tiếp vào ô (triệu chứng, dấu hiệu cảnh báo, mức độ, lời khuyên...).
  - Đánh dấu trạng thái duyệt + ghi chú.

Sau khi bác sĩ sửa xong, dùng `import_doctor_review.py` để nạp ngược các chỉnh
sửa vào JSON gốc (round-trip).

Cách chạy (từ thư mục gốc dự án):
    python scripts/qa/export_for_doctor_review.py

Tùy chọn:
    --out PATH                  Đường dẫn file xlsx xuất ra
    --diseases-limit N          Giới hạn số bệnh (0 = tất cả). Mặc định 0.
    --interactions-limit N      Giới hạn số tương tác. Mặc định 1500.
    --interactions-min-severity {low,medium,high}
                                Chỉ xuất tương tác >= mức này. Mặc định medium.

Cột "ID (không sửa)" là khóa để nạp ngược — KHÔNG được xóa/sửa cột này.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
KB = ROOT / "data" / "knowledge_base"

DISEASE_FILES = [
    KB / "vietnam_common_diseases.json",
    KB / "vietnam_diseases_full.json",
]
INTERACTION_FILE = KB / "drug_interactions.json"

SEVERITY_ORDER = {"high": 0, "medium": 1, "low": 2, "": 3, None: 3}
SEVERITY_VI = {"high": "Nặng (high)", "medium": "Trung bình (medium)", "low": "Nhẹ (low)"}

# ── Styles ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # vàng nhạt cho cột bác sĩ điền
REVIEW_HEADER_FILL = PatternFill("solid", fgColor="BF8F00")
LOCK_FILL = PatternFill("solid", fgColor="D9D9D9")  # xám cho cột ID khóa
SEV_FILL = {
    "high": PatternFill("solid", fgColor="F4CCCC"),     # đỏ nhạt
    "medium": PatternFill("solid", fgColor="FCE5CD"),   # cam nhạt
    "low": PatternFill("solid", fgColor="D9EAD3"),      # xanh nhạt
}
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── JSON loading (streaming cho file lớn) ────────────────────────────────────
def stream_json_array(path: Path) -> Iterable[dict[str, Any]]:
    """Đọc lần lượt từng object trong JSON array khổng lồ (tránh nạp cả file)."""
    with path.open("r", encoding="utf-8", errors="replace") as f:
        buf = ""
        depth = 0
        in_str = False
        esc = False
        start = -1
        seen = False
        while True:
            chunk = f.read(262144)
            if not chunk:
                break
            buf += chunk
            i = 0
            while i < len(buf):
                c = buf[i]
                if not seen:
                    if c == "[":
                        seen = True
                    i += 1
                    continue
                if esc:
                    esc = False
                elif c == "\\":
                    esc = True
                elif c == '"':
                    in_str = not in_str
                elif not in_str:
                    if c == "{":
                        if depth == 0:
                            start = i
                        depth += 1
                    elif c == "}":
                        depth -= 1
                        if depth == 0 and start >= 0:
                            try:
                                yield json.loads(buf[start : i + 1])
                            except json.JSONDecodeError:
                                pass
                            start = -1
                i += 1
            if depth == 0 and start < 0:
                buf = buf[i:]
            elif start > 0:
                buf = buf[start:]
                start = 0


def load_json_array(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [x for x in json.loads(path.read_text(encoding="utf-8")) if isinstance(x, dict)]


def _join(values: Any) -> str:
    if isinstance(values, list):
        return "\n".join(str(v).strip() for v in values if str(v).strip())
    if values is None:
        return ""
    return str(values).strip()


# ── Sheet builders ───────────────────────────────────────────────────────────
def _style_header(ws, headers: list[str], review_cols: set[str], lock_cols: set[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        if name in review_cols:
            cell.fill = REVIEW_HEADER_FILL
        elif name in lock_cols:
            cell.fill = PatternFill("solid", fgColor="808080")
        else:
            cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _write_row(ws, row_idx: int, headers: list[str], values: dict[str, Any],
               review_cols: set[str], lock_cols: set[str], severity: str | None) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=values.get(name, ""))
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if name in review_cols:
            cell.fill = REVIEW_FILL
        elif name in lock_cols:
            cell.fill = LOCK_FILL
        elif name == "Mức độ (sửa được)" and severity in SEV_FILL:
            cell.fill = SEV_FILL[severity]


def _set_widths(ws, widths: dict[str, int], headers: list[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)


def build_diseases_sheet(ws, diseases: list[dict[str, Any]]) -> int:
    headers = [
        "ID (không sửa)",
        "Tên bệnh",
        "Mã ICD-10",
        "Phân loại",
        "Triệu chứng thường gặp (sửa được)",
        "Dấu hiệu cần khám gấp / red flags (sửa được)",
        "Biến chứng (sửa được)",
        "Mức độ (sửa được)",
        "Lời khuyên / xử trí (sửa được)",
        "Mô tả đầy đủ (sửa được)",
        "Nguồn",
        "Độ tin cậy hiện tại",
        "✅ Trạng thái duyệt",
        "📝 Ghi chú của bác sĩ",
    ]
    review_cols = {
        "Triệu chứng thường gặp (sửa được)",
        "Dấu hiệu cần khám gấp / red flags (sửa được)",
        "Biến chứng (sửa được)",
        "Mức độ (sửa được)",
        "Lời khuyên / xử trí (sửa được)",
        "Mô tả đầy đủ (sửa được)",
        "✅ Trạng thái duyệt",
        "📝 Ghi chú của bác sĩ",
    }
    lock_cols = {"ID (không sửa)"}
    widths = {
        "ID (không sửa)": 26, "Tên bệnh": 34, "Mã ICD-10": 12, "Phân loại": 20,
        "Triệu chứng thường gặp (sửa được)": 34,
        "Dấu hiệu cần khám gấp / red flags (sửa được)": 34,
        "Biến chứng (sửa được)": 26, "Mức độ (sửa được)": 18,
        "Lời khuyên / xử trí (sửa được)": 38, "Mô tả đầy đủ (sửa được)": 50,
        "Nguồn": 24, "Độ tin cậy hiện tại": 14,
        "✅ Trạng thái duyệt": 20, "📝 Ghi chú của bác sĩ": 40,
    }
    _style_header(ws, headers, review_cols, lock_cols)
    _set_widths(ws, widths, headers)

    row_idx = 2
    for d in diseases:
        s = d.get("structured") or {}
        sev = (s.get("severity") or "").strip().lower() or None
        symptoms = s.get("common_symptoms") or s.get("symptoms") or []
        values = {
            "ID (không sửa)": d.get("id", ""),
            "Tên bệnh": s.get("name") or d.get("title", ""),
            "Mã ICD-10": s.get("icd10_code", ""),
            "Phân loại": s.get("category", ""),
            "Triệu chứng thường gặp (sửa được)": _join(symptoms),
            "Dấu hiệu cần khám gấp / red flags (sửa được)": _join(s.get("red_flags")),
            "Biến chứng (sửa được)": _join(s.get("common_complications")),
            "Mức độ (sửa được)": SEVERITY_VI.get(sev, sev or ""),
            "Lời khuyên / xử trí (sửa được)": _join(s.get("advice")),
            "Mô tả đầy đủ (sửa được)": d.get("content", ""),
            "Nguồn": (d.get("source") or {}).get("name", ""),
            "Độ tin cậy hiện tại": d.get("confidence", ""),
            "✅ Trạng thái duyệt": "",
            "📝 Ghi chú của bác sĩ": "",
        }
        _write_row(ws, row_idx, headers, values, review_cols, lock_cols, sev)
        row_idx += 1

    _add_status_validation(ws, headers, "✅ Trạng thái duyệt", row_idx)
    _add_severity_validation(ws, headers, "Mức độ (sửa được)", row_idx)
    return row_idx - 2


def build_interactions_sheet(ws, interactions: list[dict[str, Any]]) -> int:
    headers = [
        "ID (không sửa)",
        "Thuốc A",
        "Thuốc B",
        "Mức độ (sửa được)",
        "Cơ chế tương tác (sửa được)",
        "Khuyến nghị xử trí (sửa được)",
        "Mô tả đầy đủ (sửa được)",
        "Nguồn",
        "Độ tin cậy hiện tại",
        "✅ Trạng thái duyệt",
        "📝 Ghi chú của bác sĩ",
    ]
    review_cols = {
        "Mức độ (sửa được)", "Cơ chế tương tác (sửa được)",
        "Khuyến nghị xử trí (sửa được)", "Mô tả đầy đủ (sửa được)",
        "✅ Trạng thái duyệt", "📝 Ghi chú của bác sĩ",
    }
    lock_cols = {"ID (không sửa)"}
    widths = {
        "ID (không sửa)": 20, "Thuốc A": 24, "Thuốc B": 24, "Mức độ (sửa được)": 18,
        "Cơ chế tương tác (sửa được)": 46, "Khuyến nghị xử trí (sửa được)": 44,
        "Mô tả đầy đủ (sửa được)": 50, "Nguồn": 28, "Độ tin cậy hiện tại": 14,
        "✅ Trạng thái duyệt": 20, "📝 Ghi chú của bác sĩ": 40,
    }
    _style_header(ws, headers, review_cols, lock_cols)
    _set_widths(ws, widths, headers)

    row_idx = 2
    for it in interactions:
        s = it.get("structured") or {}
        sev = (s.get("severity") or "").strip().lower() or None
        values = {
            "ID (không sửa)": it.get("id", ""),
            "Thuốc A": s.get("drug_a", ""),
            "Thuốc B": s.get("drug_b", ""),
            "Mức độ (sửa được)": SEVERITY_VI.get(sev, sev or ""),
            "Cơ chế tương tác (sửa được)": s.get("mechanism", ""),
            "Khuyến nghị xử trí (sửa được)": s.get("recommendation", ""),
            "Mô tả đầy đủ (sửa được)": it.get("content", ""),
            "Nguồn": (it.get("source") or {}).get("name", ""),
            "Độ tin cậy hiện tại": it.get("confidence", ""),
            "✅ Trạng thái duyệt": "",
            "📝 Ghi chú của bác sĩ": "",
        }
        _write_row(ws, row_idx, headers, values, review_cols, lock_cols, sev)
        row_idx += 1

    _add_status_validation(ws, headers, "✅ Trạng thái duyệt", row_idx)
    _add_severity_validation(ws, headers, "Mức độ (sửa được)", row_idx)
    return row_idx - 2


def _add_status_validation(ws, headers: list[str], col_name: str, last_row: int) -> None:
    col = get_column_letter(headers.index(col_name) + 1)
    dv = DataValidation(
        type="list",
        formula1='"Chưa duyệt,Đã duyệt - đúng,Đã sửa,Cần xóa,Cần xem lại"',
        allow_blank=True,
    )
    dv.prompt = "Chọn trạng thái duyệt"
    dv.promptTitle = "Trạng thái"
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max(last_row, 2)}")


def _add_severity_validation(ws, headers: list[str], col_name: str, last_row: int) -> None:
    col = get_column_letter(headers.index(col_name) + 1)
    dv = DataValidation(
        type="list",
        formula1='"Nặng (high),Trung bình (medium),Nhẹ (low)"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max(last_row, 2)}")


def build_guide_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    lines = [
        ("HƯỚNG DẪN RÀ SOÁT DỮ LIỆU Y TẾ — MediSign AI", True),
        ("", False),
        ("Mục đích: Bác sĩ rà soát và chỉnh sửa dữ liệu bệnh + tương tác thuốc trước khi đưa vào sử dụng.", False),
        ("", False),
        ("CÁCH LÀM:", True),
        ("1. Mở 2 sheet: 'Bệnh' và 'Tương tác thuốc'.", False),
        ("2. Các cột nền VÀNG là cột bác sĩ được phép sửa trực tiếp.", False),
        ("3. Cột 'ID (không sửa)' nền XÁM là khóa hệ thống — KHÔNG xóa/sửa cột này.", False),
        ("4. Sửa nội dung sai ngay trong ô. Nhiều mục cách nhau bằng XUỐNG DÒNG (Alt+Enter).", False),
        ("5. Cột 'Trạng thái duyệt' chọn từ danh sách: Chưa duyệt / Đã duyệt - đúng / Đã sửa / Cần xóa / Cần xem lại.", False),
        ("6. Cột 'Ghi chú của bác sĩ' để ghi lý do sửa, nguồn tham khảo, cảnh báo...", False),
        ("", False),
        ("QUY ƯỚC MÀU MỨC ĐỘ (cột 'Mức độ'):", True),
        ("   Đỏ = Nặng (high)   |   Cam = Trung bình (medium)   |   Xanh = Nhẹ (low)", False),
        ("", False),
        ("ƯU TIÊN RÀ SOÁT:", True),
        ("- Ưu tiên các tương tác thuốc mức 'Nặng (high)' và bệnh có 'red flags'.", False),
        ("- Kiểm tra dấu hiệu cảnh báo (red flags) vì đây là phần ảnh hưởng an toàn trực tiếp.", False),
        ("", False),
        ("Sau khi sửa xong: lưu file và gửi lại cho nhóm kỹ thuật để nạp ngược vào hệ thống.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120


def main() -> None:
    ap = argparse.ArgumentParser(description="Xuất dữ liệu y tế cho bác sĩ rà soát (.xlsx)")
    ap.add_argument("--out", default=str(ROOT / "data" / "review" / "medisign_doctor_review.xlsx"))
    ap.add_argument("--diseases-limit", type=int, default=0, help="0 = tất cả")
    ap.add_argument("--interactions-limit", type=int, default=1500)
    ap.add_argument(
        "--interactions-min-severity",
        choices=["low", "medium", "high"],
        default="medium",
    )
    args = ap.parse_args()

    # ── Load diseases ────────────────────────────────────────────────────
    diseases: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for path in DISEASE_FILES:
        for d in load_json_array(path):
            did = d.get("id")
            if did and did in seen_ids:
                continue
            if did:
                seen_ids.add(did)
            diseases.append(d)
    # Sắp xếp: bệnh có red_flags / severity cao lên trước
    diseases.sort(
        key=lambda d: (
            SEVERITY_ORDER.get(((d.get("structured") or {}).get("severity") or "").lower(), 3),
            0 if (d.get("structured") or {}).get("red_flags") else 1,
        )
    )
    if args.diseases_limit > 0:
        diseases = diseases[: args.diseases_limit]

    # ── Load interactions (streaming + lọc theo mức độ) ──────────────────
    min_sev_rank = SEVERITY_ORDER[args.interactions_min_severity]
    interactions: list[dict[str, Any]] = []
    if INTERACTION_FILE.exists():
        for it in stream_json_array(INTERACTION_FILE):
            sev = ((it.get("structured") or {}).get("severity") or "").lower()
            if SEVERITY_ORDER.get(sev, 3) <= min_sev_rank:
                interactions.append(it)
                if args.interactions_limit > 0 and len(interactions) >= args.interactions_limit:
                    break
    interactions.sort(
        key=lambda it: SEVERITY_ORDER.get(((it.get("structured") or {}).get("severity") or "").lower(), 3)
    )

    # ── Build workbook ───────────────────────────────────────────────────
    wb = Workbook()
    guide = wb.active
    guide.title = "Hướng dẫn"
    build_guide_sheet(guide)

    ws_d = wb.create_sheet("Bệnh")
    n_d = build_diseases_sheet(ws_d, diseases)

    ws_i = wb.create_sheet("Tương tác thuốc")
    n_i = build_interactions_sheet(ws_i, interactions)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print("✅ Đã xuất file rà soát cho bác sĩ:")
    print(f"   {out_path}")
    print(f"   - Sheet 'Bệnh': {n_d:,} dòng")
    print(f"   - Sheet 'Tương tác thuốc': {n_i:,} dòng (mức >= {args.interactions_min_severity})")
    print("\nGửi file này cho bác sĩ. Sau khi sửa xong, dùng import_doctor_review.py để nạp ngược.")


if __name__ == "__main__":
    main()
